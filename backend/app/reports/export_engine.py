import io
import csv
from typing import List, Dict, Any, Optional
from datetime import datetime


def generate_csv_bytes(data_list: List[Dict[str, Any]], headers_map: Dict[str, str]) -> bytes:
    """
    توليد ملف CSV مع إضافة UTF-8 BOM (\xef\xbb\xbf) لفتح اللغة العربية تلقائياً وبدقة ببرنامج Excel
    """
    output = io.StringIO()
    # Add UTF-8 BOM for Arabic support in Excel
    output.write('\ufeff')

    writer = csv.writer(output)
    
    # Write Headers
    header_keys = list(headers_map.keys())
    header_labels = list(headers_map.values())
    writer.writerow(header_labels)

    # Write Rows
    for row in data_list:
        line = [str(row.get(k, '')) for k in header_keys]
        writer.writerow(line)

    return output.getvalue().encode('utf-8')


def generate_excel_bytes(data_list: List[Dict[str, Any]], headers_map: Dict[str, str], title: str) -> bytes:
    """
    توليد ملف Excel (.xlsx) منسق بخلايا ملونة باللغة العربية RTL
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:30]
        ws.views.sheetView[0].rightToLeft = True # RTL

        # Styles
        header_fill = PatternFill(start_color="7A081D", end_color="7A081D", fill_type="solid") # Church Maroon
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell_font = Font(name="Segoe UI", size=10)
        align_right = Alignment(horizontal="right", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style='thin', color='D4AF37'),
            right=Side(style='thin', color='D4AF37'),
            top=Side(style='thin', color='D4AF37'),
            bottom=Side(style='thin', color='D4AF37')
        )

        # Title Row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers_map))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(name="Segoe UI", size=14, bold=True, color="7A081D")
        title_cell.alignment = align_center

        # Header Row
        header_keys = list(headers_map.keys())
        for col_idx, k in enumerate(header_keys, 1):
            cell = ws.cell(row=3, column=col_idx, value=headers_map[k])
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = thin_border

        # Data Rows
        for row_idx, row_data in enumerate(data_list, 4):
            for col_idx, k in enumerate(header_keys, 1):
                val = row_data.get(k, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = cell_font
                cell.alignment = align_right
                cell.border = thin_border

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 14)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except ImportError:
        # Fallback to CSV if openpyxl not installed
        return generate_csv_bytes(data_list, headers_map)


def generate_pdf_html(
    data_list: List[Dict[str, Any]],
    headers_map: Dict[str, str],
    title: str,
    church_name: str,
    summary_cards: Optional[List[Dict[str, Any]]] = None
) -> str:
    """
    توليد وثيقة HTML/PDF مروسة رسمياً بشعار واسم الكنيسة مخصصة للطباعة والتصدير
    """
    now_str = datetime.now().strftime("%Y/%m/%d - %I:%M %p")
    header_keys = list(headers_map.keys())

    cards_html = ""
    if summary_cards:
        cards_html = '<div style="display: flex; gap: 15px; margin-bottom: 20px; flex-wrap: wrap;">'
        for c in summary_cards:
            cards_html += f'''
            <div style="flex: 1; min-width: 180px; padding: 12px 15px; background: #fff8f8; border: 1px solid #d4af37; border-radius: 8px; text-align: center;">
                <div style="font-size: 12px; color: #7a081d; font-weight: bold;">{c.get("label")}</div>
                <div style="font-size: 20px; font-weight: 900; color: #3b000b; margin-top: 4px;">{c.get("value")}</div>
            </div>
            '''
        cards_html += '</div>'

    rows_html = ""
    for idx, row in enumerate(data_list, 1):
        rows_html += '<tr>'
        for k in header_keys:
            val = row.get(k, '')
            rows_html += f'<td style="padding: 8px 12px; border-bottom: 1px solid #eee; text-align: right;">{val}</td>'
        rows_html += '</tr>'

    headers_html = "".join([f'<th style="padding: 10px 12px; background: #7a081d; color: #fbeea9; text-align: right;">{headers_map[k]}</th>' for k in header_keys])

    html_content = f'''<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head>
    <meta charset="utf-8">
    <title>{title}</title>
    <style>
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #fff; color: #222; margin: 0; padding: 20px; }}
        .header-box {{ display: flex; align-items: center; justify-content: space-between; border-bottom: 3px double #d4af37; padding-bottom: 15px; margin-bottom: 20px; }}
        .logo-title {{ display: flex; alignItems: center; gap: 15px; }}
        .title-main {{ font-size: 22px; font-weight: 900; color: #7a081d; margin: 0; }}
        .sub-title {{ font-size: 13px; color: #666; margin-top: 3px; }}
        .date-box {{ font-size: 11px; color: #888; text-align: left; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 15px; font-size: 13px; }}
        .footer-note {{ margin-top: 30px; font-size: 11px; color: #999; border-top: 1px solid #ddd; padding-top: 10px; text-align: center; }}
        @media print {{
            body {{ padding: 0; }}
            .no-print {{ display: none; }}
        }}
    </style>
</head>
<body>
    <div class="no-print" style="margin-bottom: 15px; text-align: left;">
        <button onclick="window.print()" style="padding: 8px 18px; background: #7a081d; color: #fff; border: none; border-radius: 6px; cursor: pointer; font-weight: bold;">
            طباعة التقرير / حفظ PDF 🖨️
        </button>
    </div>

    <div class="header-box">
        <div class="logo-title">
            <div>
                <h1 class="title-main">{church_name}</h1>
                <div class="sub-title">نظام المسلة الكنسي المركزي — {title}</div>
            </div>
        </div>
        <div class="date-box">
            <div>تاريخ التصدير: {now_str}</div>
            <div>تقرير رسمي معتمد</div>
        </div>
    </div>

    {cards_html}

    <table>
        <thead>
            <tr>{headers_html}</tr>
        </thead>
        <tbody>
            {rows_html}
        </tbody>
    </table>

    <div class="footer-note">
        تم استخراج هذا التقرير آلياً من نظام المسلة — كنيسة السيدة العذراء مريم والأنبا بولا بالمسلة (مطرانية أسوان).
    </div>
</body>
</html>'''
    return html_content
